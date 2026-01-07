from __future__ import annotations

from pathlib import Path
from typing import Any
from itertools import chain

from pydantic import BaseModel, ConfigDict, Field, model_validator, validate_call, PositiveFloat
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from .PipetG import PipetG

# --- Table Design Constants ---
THIN_SIDE = Side(style="thin")
GRID_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
# --- Colors ---
LIGHT_ORANGE = PatternFill("solid", fgColor="FFF2CC")
LIGHT_BLUE = PatternFill("solid", fgColor="DDEBF7")
WHITE = PatternFill("solid", fgColor="FFFFFF")
HEADER_ORANGE = PatternFill("solid", fgColor="FFA62B")
VIAL_IDX_FILL = PatternFill("solid", fgColor="16697A")
HEADER_BOLD = Font(bold=True)


class InputXlsx(BaseModel):
    pipet: PipetG

    xlsx_path: Path | None = Field(default=None, exclude=True, repr=False)
    wb: Workbook | None = Field(default=None, exclude=True, repr=False)

    model_config = ConfigDict(arbitrary_types_allowed=True, validate_assignment=True)

    # ---------- template creation ----------
    def create_empty_table(self, out_path: Path, *, stages: int = 3) -> Path:
        wb = Workbook()
        wb.remove(wb.active)

        vial_start_idx = 1
        solvent_start_idx = 1

        for rack in self.pipet.setup.racks:
            solvent_start_idx = self._write_solvent_sheet(wb, rack, solvent_start_idx)
            vial_start_idx = self._write_vial_sheet(wb, rack, vial_start_idx, stages=stages)

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def _write_solvent_sheet(self, wb: Workbook, rack, start_index: int) -> int:
        ws = wb.create_sheet(f"{rack.name}_solvents")

        n_rows = int(rack.solvent_rows)
        n_cols = int(rack.solvent_columns)

        # make the rest of the visible region white
        for row in ws.iter_rows(min_row=1, max_row=n_rows + 40, min_col=1, max_col=37):
            for cell in row:
                cell.fill = WHITE

        instructions = [
            "1) Hover over the colored grid to see the solvent index.",
            "2) Fill in the solvent id.",
            "",
            "Without a filled in solvent id, the solvent is not valid. If you do not need precise pipetting, "
            "use the solvent id of a hydrodynamically similar or generic solvent, such as water or acetone.",
            "",
            "<<<---Where to find solvent id?-->>>",
            "Inside config/DB/liquid_handling.db find the solvent list in the solvent table.",
            "Any solvent can be used with the selected syringe only if they have a couple listed inside the "
            "syringe solvent link table where the calibrated pipeting parameters are stored for the software.",
        ]
        for i, line in enumerate(instructions):
            cell = ws.cell(row=n_rows + 2 + i, column=10, value=line)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.font = HEADER_BOLD
            if i == 5:
                cell.font = Font(color="FFA62B", bold=True)

        idx = start_index
        for col in range(1, n_cols + 1):
            for row in range(1, n_rows + 1):
                cell = ws.cell(row=row, column=col, value=None)  # user writes solvent_id here

                cell.fill = LIGHT_ORANGE
                cell.border = GRID_BORDER
                cell.alignment = CENTER_ALIGN

                cell.comment = Comment(f"{idx}", "index") # comment shows global solvent-slot index
                idx += 1

        return start_index + (n_rows * n_cols)

    def _write_vial_sheet(self, wb: Workbook, rack, start_index: int, *, stages: int = 3) -> int:
        ws = wb.create_sheet(f"{rack.name}_vials")

        n_rows = int(rack.vial_rows)
        n_cols = int(rack.vial_columns)
        n = n_rows * n_cols # total number of vials

        # make the rest of the visible region white
        table_top = n_rows + 2
        header_max_col = 1 + (3 * stages)

        rows = chain(ws.iter_rows(min_row=1, max_row=n_rows + 1, min_col=1, max_col=37),
            ws.iter_rows(min_row=table_top + n + 1, max_row=table_top + n + 30, min_col=1, max_col=37))
        for row in rows:
            for cell in row:
                cell.fill = WHITE

        # --- vial grid with global vial indices ---
        idx = start_index
        for col in range(1, n_cols + 1):
            for row in range(1, n_rows + 1):
                cell = ws.cell(row=row, column=col, value=idx)
                cell.fill = LIGHT_BLUE
                cell.border = GRID_BORDER
                cell.alignment = CENTER_ALIGN
                idx += 1

        # --- the program table below: A=index, then (volume_uL, solvent_index, flush)*stages ---

        # headers for stages
        for col in ws.iter_cols(min_row=table_top, max_row=table_top, min_col=1, max_col=header_max_col):
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = HEADER_ORANGE
                cell.font = HEADER_BOLD
        ws.row_dimensions[table_top].height = 30

        hint_cell = ws.cell(row=table_top - 1, column=10, value="Find the instruction below the vial program.")
        hint_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        hint_cell.font = HEADER_BOLD

        ws.cell(row=table_top, column=1, value="index") # the vial index header
        for i in range(n):
            r = table_top + 1 + i
            cell = ws.cell(row=r, column=1, value=start_index + i)
            cell.fill = VIAL_IDX_FILL
            cell.font = Font(color="FFFFFF")

        for col in range(2, 1 + header_max_col, 3): # range(start, end, stepsize)
            ws.cell(row=table_top, column=col + 0, value="volume\nµL")
            ws.cell(row=table_top, column=col + 1, value="solvent\nindex")

            ws.cell(row=table_top, column=col + 2, value="flush")
            for i in range(n): # FALSE in every Flush cell
                r = table_top + 1 + i
                ws.cell(row=r, column=col + 2, value=False)

        # rows and borders for the program table
        for row in ws.iter_rows(min_row=table_top, max_row=table_top + n, min_col=1, max_col=header_max_col):
            for cell in row:
                cell.border = GRID_BORDER

        instructions = [
            "1) Enter volume (µL) and solvent index for each stage.",
            "",
            "<<<---What is a stage?-->>>",
            "",
            "Stage is a group of columns: {volume µL, solvent index} or {volume µL, solvent index, flush}.}",
            "The pipetting program is generated stage by stage (downwards).",
            "Add as many stages as you need by copy pasting columns in the table.",
            "",
            "<<<---Flushing the syringe--->>>"
            "2) Optional: enter flush TRUE to flush with the solvent used pipetting "
            "or an integer solvent index to flush with a specific solvent.",
            "",
            "3) Fill the cell containing a volume value with any color "
            "to dispense the liquid slowly (crystallization?).",
            "",
            "4) Leave a stage row empty to skip it.",
        ]
        for i, line in enumerate(instructions):
            cell = ws.cell(row=table_top + n + 2 + i, column=10, value=line)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.font = HEADER_BOLD
            if i == 2:
                cell.font = Font(color="FFA62B", bold=True)

        return start_index + n

    # ------------------------------------------------------------------
    # Loading / validation
    # ------------------------------------------------------------------

    def load(self, xlsx_path: str | Path, **wb_kwargs: Any) -> "InputXlsx":
        """Load an existing Excel file into this instance.

        Assigning `self.wb` triggers model validators (validate_assignment=True).
        """
        path = Path(xlsx_path)
        self.xlsx_path = path
        self.wb = load_workbook(filename=str(path), **wb_kwargs)
        return self

    @classmethod
    def from_excel(cls, xlsx_path: str | Path, pipet: PipetG, **wb_kwargs: Any) -> "InputXlsx":
        obj = cls(pipet=pipet)
        obj.load(xlsx_path, **wb_kwargs)  # triggers validation
        return obj

    @classmethod
    def write_gcode_from_excel(
        cls,
        xlsx_path: str | Path,
        pipet: PipetG,
        **g_kwargs: Any,
    ) -> Path:
        obj = cls.from_excel(xlsx_path, pipet=pipet)
        return obj.generate_gcode(**g_kwargs)

    @model_validator(mode="after")
    def _validate_solvent_ids_in_solvent_grids(self) -> "InputXlsx":
        if self.wb is None:
            return self
        self._validate_solvent_ids_in_solvent_grids_impl()
        return self

    @model_validator(mode="after")
    def _validate_solvent_index_in_vial_program(self) -> "InputXlsx":
        if self.wb is None:
            return self
        self._validate_vial_program_bounds_impl()
        return self

    # ---------------- validation helpers ----------------

    def _is_empty(self, v: Any) -> bool:
        return v is None or (isinstance(v, str) and v.strip() == "")

    def _as_int(self, v: Any, *, where: str) -> int | None:
        if v is None:
            return None
        if isinstance(v, bool):
            raise ValueError(f"{where}: boolean is not a valid integer")
        if isinstance(v, int):
            return v
        if isinstance(v, float):
            if v.is_integer():
                return int(v)
            raise ValueError(f"{where}: expected integer, got float {v}")
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                return None
            if "," in s and "." not in s: # regional formatting 3,0 -> 3.0
                s = s.replace(",", ".")
            try:
                f = float(s)
            except ValueError as e:
                raise ValueError(f"{where}: cannot parse '{v}' as number") from e
            if f.is_integer():
                return int(f)
            raise ValueError(f"{where}: expected integer, got '{v}'")
        raise ValueError(f"{where}: unsupported type {type(v).__name__}")

    def _parse_flush_spec(self, v: Any, *, where: str) -> bool | int | None:
        """Flush cell can be:
        - empty -> None (no flush)
        - TRUE/FALSE -> bool
        - integer k:
            - 0 -> False (no flush)
            - k>=1 -> flush with global solvent_index=k
        """
        if v is None:
            return None

        if isinstance(v, bool):
            return v

        def for_int(val: int):
            if val == 0:
                return False
            if val >= 1:
                return val
            raise ValueError(f"{where}: flush integer must be >=0, got {val}")

        if isinstance(v, int):
            return for_int(v)

        if isinstance(v, float):
            if not v.is_integer():
                raise ValueError(f"{where}: flush must be TRUE/FALSE or an integer solvent index, got {v}")
            iv = int(v)
            return for_int(iv)

        if isinstance(v, str):
            s = v.strip()
            if s == "":
                return None
            if "," in s and "." not in s: # regional formatting 3,0 -> 3.0
                s = s.replace(",", ".")
            s_low = s.lower()
            if s_low in {"true", "t", "yes", "y"}:
                return True
            if s_low in {"false", "f", "no", "n"}:
                return False

            try:
                f = float(s_low)
            except ValueError as e:
                raise ValueError(f"{where}: flush cannot parse '{v}' (use TRUE/FALSE or a solvent index)") from e
            if not f.is_integer():
                raise ValueError(f"{where}: flush must be TRUE/FALSE or integer solvent index, got '{v}'")
            iv = int(f)
            return for_int(iv)

        raise ValueError(f"{where}: unsupported type {type(v).__name__} for flush")

    def _allowed_solvent_ids(self) -> set[int]:
        ids = []
        for s in self.pipet.setup.solvents:
            if s.id is None:
                raise ValueError("Setup contains a Solvent with id=None; cannot validate solvent IDs.")
            ids.append(int(s.id))
        return set(ids)

    def _total_solvent_slots(self) -> int:
        return sum(r.solvent_rows * r.solvent_columns for r in self.pipet.setup.racks)

    # ---------- program table parsing (multi-stage) ----------

    def _program_header_last_col(self, ws, *, header_row: int) -> int:
        """Last column (in header_row) that has a non-empty header value."""
        last = 1
        max_col = ws.max_column or 1
        for col in range(1, max_col + 1):
            if not self._is_empty(ws.cell(row=header_row, column=col).value):
                last = col
        return last

    def _parse_program_stages(self, ws, *, sheet_name: str, header_row: int) -> list[tuple[int, int, int | None]]:
        """Return list of stages as tuples: (vol_col, solvent_col, flush_col|None)

        Expected layout (after col A=index):
            [volume_uL, solvent_index] OR [volume_uL, solvent_index, flush]
        repeated as many times as the user adds.
        """
        last_col = self._program_header_last_col(ws, header_row=header_row)
        if last_col < 3:
            raise ValueError(f"{sheet_name}: program header row {header_row} is missing required columns.")

        def norm(x: Any) -> str:
            if x is None:
                return ""
            return str(x).strip().lower().replace(" ", "")

        def is_volume(h: str) -> bool:
            return h.startswith("volume")

        def is_solvent(h: str) -> bool:
            return h.startswith("solvent")

        def is_flush(h: str) -> bool:
            return h.startswith("flush")

        stages: list[tuple[int, int, int | None]] = []
        col = 2  # after index
        while col <= last_col:
            h_vol = norm(ws.cell(row=header_row, column=col).value)
            h_sol = norm(ws.cell(row=header_row, column=col + 1).value) if (col + 1) <= last_col else ""

            # stop when the user didn't define more stages
            if h_vol == "" and h_sol == "":
                break

            if not is_volume(h_vol):
                raise ValueError(
                    f"{sheet_name}!{ws.cell(row=header_row, column=col).coordinate}: "
                    f"expected a volume header (e.g. 'volume_uL'), got '{ws.cell(row=header_row, column=col).value}'."
                )
            if not is_solvent(h_sol):
                raise ValueError(
                    f"{sheet_name}!{ws.cell(row=header_row, column=col+1).coordinate}: "
                    f"expected a solvent header (e.g. 'solvent_index'), got '{ws.cell(row=header_row, column=col+1).value}'."
                )

            vol_col = col
            sol_col = col + 1
            flush_col: int | None = None
            # check for the flush column
            if (col + 2) <= last_col:
                h_f = norm(ws.cell(row=header_row, column=col + 2).value)
                if is_flush(h_f):
                    flush_col = col + 2
                    col += 3
                else:
                    col += 2
            else:
                col += 2

            stages.append((vol_col, sol_col, flush_col))

        if not stages:
            raise ValueError(f"{sheet_name}: no program stages found in header row {header_row}.")
        return stages

    # ---------------- validation impl ----------------

    def _validate_solvent_ids_in_solvent_grids_impl(self) -> None:
        assert self.wb is not None
        allowed = self._allowed_solvent_ids()

        for rack in self.pipet.setup.racks:
            name = f"{rack.name}_solvents"
            if name not in self.wb.sheetnames:
                raise ValueError(f"Missing sheet '{name}' in workbook.")

            ws = self.wb[name]
            n_rows = int(rack.solvent_rows)
            n_cols = int(rack.solvent_columns)

            for col in range(1, n_cols + 1):
                for row in range(1, n_rows + 1):
                    cell = ws.cell(row=row, column=col)
                    s_id = self._as_int(cell.value, where=f"{name}!{cell.coordinate}")
                    if s_id is None:
                        continue
                    if s_id not in allowed:
                        raise ValueError(
                            f"{name}!{cell.coordinate}: solvent_id={s_id} not in setup.solvents ids={sorted(allowed)}"
                        )

    def _validate_vial_program_bounds_impl(self) -> None:
        """Validate solvent_index (all stages) + flush integer bounds (all stages)."""
        assert self.wb is not None
        total_slots = self._total_solvent_slots()

        for rack in self.pipet.setup.racks:
            name = f"{rack.name}_vials"
            if name not in self.wb.sheetnames:
                raise ValueError(f"Missing sheet '{name}' in workbook.")

            ws = self.wb[name]
            n_rows = int(rack.vial_rows)
            n_cols = int(rack.vial_columns)
            n = n_rows * n_cols

            header_row = n_rows + 2
            stages = self._parse_program_stages(ws, sheet_name=name, header_row=header_row)

            for i in range(n):
                r = header_row + 1 + i

                for (vol_col, sol_col, flush_col) in stages:
                    sol_cell = ws.cell(row=r, column=sol_col)
                    s_idx = self._as_int(sol_cell.value, where=f"{name}!{sol_cell.coordinate}")
                    if s_idx is not None:
                        if not (1 <= s_idx <= total_slots):
                            raise ValueError(
                                f"{name}!{sol_cell.coordinate}: solvent_index={s_idx} "
                                f"out of bounds (allowed 1..{total_slots})"
                            )
                    # if only flush is filled, it will still raise an error.
                    if flush_col is not None:
                        f_cell = ws.cell(row=r, column=flush_col)
                        spec = self._parse_flush_spec(f_cell.value, where=f"{name}!{f_cell.coordinate}")
                        if isinstance(spec, bool):
                            continue
                        if isinstance(spec, int):
                            if not (1 <= spec <= total_slots):
                                raise ValueError(
                                    f"{name}!{f_cell.coordinate}: flush solvent_index={spec} "
                                    f"out of bounds (allowed 1..{total_slots})"
                                )

    # ------------------------------------------------------------------
    # G-code generation
    # ------------------------------------------------------------------

    def _solvent_id_by_global_solvent_index(self) -> list[int | None]:
        """mapping[solvent_index-1] -> solvent_id or None (if not assigned in grid)."""
        assert self.wb is not None

        mapping: list[int | None] = []
        for rack in self.pipet.setup.racks:
            ws = self.wb[f"{rack.name}_solvents"]
            n_rows = int(rack.solvent_rows)
            n_cols = int(rack.solvent_columns)

            for col in range(1, n_cols + 1):
                for row in range(1, n_rows + 1):
                    cell = ws.cell(row=row, column=col)
                    s_id = self._as_int(cell.value, where=f"{rack.name}_solvents!{cell.coordinate}")
                    mapping.append(s_id)
        return mapping

    @validate_call
    def _split_volume_ul(self, total_ul: float, max_ul: PositiveFloat, *, eps: float = 1e-9) -> list[float]:
        """Split total volume into chunks <= max_ul."""
        if total_ul <= 0:
            raise ValueError(f"volume_uL must be > 0, got {total_ul}") # user-friendly error message

        parts: list[float] = []
        remaining = float(total_ul)

        while remaining > max_ul + eps:
            parts.append(float(max_ul))
            remaining -= float(max_ul)

        if remaining > eps:
            parts.append(float(remaining))

        return parts

    def generate_gcode(self, **g_kwargs: Any) -> Path:
        """Read the workbook program and emit gcode using the provided PipetG.

        Multi-stage handling:
          - You can add as many stage groups as you want:
              volume_uL | solvent_index | [flush]
            repeated to the right.
          - Execution is stage-wise:
              stage1 over all vials -> stage2 over all vials -> ...
        """
        if self.wb is None:
            raise RuntimeError("No workbook loaded. Use InputXlsx.from_excel(...) or .load(...) first.")

        solvent_id_map = self._solvent_id_by_global_solvent_index()

        pg = self.pipet
        pg.start(**g_kwargs)

        try:
            max_ul = float(pg.max_volume_ul)  # type: ignore[arg-type]

            pg.home()

            for rack in pg.setup.racks:
                sheet_name = f"{rack.name}_vials"
                ws = self.wb[sheet_name]
                n_rows = int(rack.vial_rows)
                n_cols = int(rack.vial_columns)
                n = n_rows * n_cols

                header_row = n_rows + 2
                stages = self._parse_program_stages(ws, sheet_name=sheet_name, header_row=header_row)

                # stage-wise execution
                for stage_i, (vol_col, sol_col, flush_col) in enumerate(stages, start=1):
                    for i in range(n):
                        r = header_row + 1 + i

                        vial_cell = ws.cell(row=r, column=1)  # vial index
                        vol_cell = ws.cell(row=r, column=vol_col)
                        s_idx_cell = ws.cell(row=r, column=sol_col)
                        flush_cell = ws.cell(row=r, column=flush_col) if flush_col is not None else None

                        vial_index = self._as_int(vial_cell.value, where=f"{sheet_name}!{vial_cell.coordinate}")
                        if vial_index is None:
                            raise ValueError(f"{sheet_name}!{vial_cell.coordinate}: vial index is empty.")

                        volume_val = vol_cell.value
                        s_idx_raw = s_idx_cell.value
                        flush_raw = flush_cell.value if flush_cell is not None else None
                        flush_spec = None
                        if flush_cell is not None:
                            flush_spec = self._parse_flush_spec(flush_raw, where=f"{sheet_name}!{flush_cell.coordinate}")

                        # skip empty stage instruction for this vial
                        if self._is_empty(volume_val) and self._is_empty(s_idx_raw) and (flush_spec in {None, False}):
                            continue

                        # flush handling (supports flush-only rows)
                        flush_idx0: int | None = None  # 0-based flush index
                        flush_solvent_id: int | None = None
                        if isinstance(flush_spec, int) and not isinstance(flush_spec, bool):
                            flush_idx0 = flush_spec - 1
                            if not (0 <= flush_idx0 < len(solvent_id_map)):
                                raise ValueError(
                                    f"{sheet_name}!{flush_cell.coordinate}: flush solvent index={flush_spec} "
                                    f"out of mapping range."
                                )
                            _s_id = solvent_id_map[flush_idx0]
                            if _s_id is None:
                                raise ValueError(
                                    f"{sheet_name}!{flush_cell.coordinate}: flush solvent index={flush_spec} "
                                    f"has no solvent_id assigned in solvent grids."
                                )
                            flush_solvent_id = int(_s_id)

                        if self._is_empty(volume_val) and self._is_empty(s_idx_raw):
                            if flush_idx0 is None or flush_solvent_id is None:
                                raise ValueError(
                                    f"{sheet_name}!{flush_cell.coordinate}: flush-only rows require an integer "
                                    f"solvent index (stage {stage_i})."
                                )
                            pg.flush(
                                volume_ul=max_ul,
                                repeats=1,
                                solvent_idx=flush_idx0,
                                solvent_id=flush_solvent_id,
                            )
                            continue

                        s_idx = self._as_int(s_idx_raw, where=f"{sheet_name}!{s_idx_cell.coordinate}")
                        if s_idx is None:
                            raise ValueError(
                                f"{sheet_name}!{s_idx_cell.coordinate}: solvent index is empty "
                                f"(stage {stage_i})."
                            )
                        if self._is_empty(volume_val):
                            raise ValueError(
                                f"{sheet_name}!{vol_cell.coordinate}: volume_uL is empty "
                                f"(stage {stage_i})."
                            )
                        # Handling volume value
                        try:
                            if isinstance(volume_val, str):
                                v_str = volume_val.strip()
                                if "," in v_str and "." not in v_str:
                                    v_str = v_str.replace(",", ".")
                                volume_ul_total = float(v_str)
                            else:
                                volume_ul_total = float(volume_val)
                        except (TypeError, ValueError) as e:
                            raise ValueError(
                                f"{sheet_name}!{vol_cell.coordinate}: volume_uL must be a number "
                                f"(stage {stage_i})."
                            ) from e
                        chunks = self._split_volume_ul(volume_ul_total, max_ul)

                        # Excel indices are 1-based; PipetG expects 0-based indices
                        vial_idx0 = vial_index - 1
                        solvent_idx0 = s_idx - 1

                        if not (0 <= solvent_idx0 < len(solvent_id_map)):
                            raise ValueError(
                                f"{sheet_name} row {r} stage {stage_i}: solvent index={s_idx} out of mapping range."
                            )

                        dispense_solvent_id = solvent_id_map[solvent_idx0]
                        if dispense_solvent_id is None:
                            raise ValueError(
                                f"{sheet_name} row {r} stage {stage_i}: solvent index={s_idx} "
                                f"has no solvent_id assigned in solvent grids."
                            )

                        if flush_spec is True:
                            flush_idx0 = solvent_idx0
                            flush_solvent_id = int(dispense_solvent_id)

                        # flush only once per stage
                        if flush_idx0 is not None and flush_solvent_id is not None:
                            flush_vol = min(volume_ul_total, max_ul) # flush no more than max_volume_ul
                            pg.flush(
                                volume_ul=flush_vol,
                                repeats=1,
                                solvent_idx=flush_idx0,
                                solvent_id=flush_solvent_id,
                            )

                        # dispense total amount (chunked)
                        slow = vol_cell.fill.fill_type is not None
                        for v_chunk in chunks:
                            pg.process_vial(
                                vial_idx=vial_idx0,
                                solvent_idx=solvent_idx0,
                                solvent_id=int(dispense_solvent_id),
                                volume_ul=v_chunk,
                                slow=slow,
                                flush_repeats=0,
                            )

            pg.finish()

            return Path(pg.outfile)

        finally:
            pg.stop()
